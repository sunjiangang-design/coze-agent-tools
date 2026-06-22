#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
伟良日报处理脚本（可复用版）
输入：主计划表 + 外购物料到货跟踪表（可选成品入库表）
输出：1. 异常报表(3个Sheet)  2. 安全库存未到货表  3. 可选：主计划表-成品入库标记
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re, os, argparse
from datetime import datetime, date

# ============ 采购关键字 ============
PROC_KW = ['筛网','铸件','圆钢','方管','圆管','卡箍','轴承','电机','螺母','螺栓',
    '尼龙网','网笼丝杆','板带','皮条','软连接','阀门','气缸','气管','弹簧',
    '超声波','电控箱','万向节','同步带','二通','卡簧','三通','平键','堵头',
    '压紧把手','装网工装','手孔盖','感应开关','筛体法兰一',
    '电源箱','磁簧式传感器','防爆电机','防爆接线盒',
    '接线管','密封垫','限位开关','节流阀','皮带轮',
    '紧固法兰','斜锥套','激振器轴','底格轴','激振器外壳','底格外壳','轴承室内套',
    '上偏心块','下偏心块','偏心块','偏心螺栓',
    '风轮','连接座','连接管','法兰',
    '上下椎板','上下锥板','托球板','筛盘','网笼','网架','进料斗','侧盖','侧盖板',
    '内压盖','筛体法兰二','活动轴套','封板',
    '电机板','轴套','大小端盖','风机法兰','管箍','液压支撑杆','投料站']

PROD_KW = ['下料','焊接','组装','粘网','机加工','机加','喷砂','喷粉','喷涂','试机','压制','包装','封箱']
IGNORE_KW = ['欠计划下单','欠下料排版','欠技术排版','欠排版']
FREEZE_PAT = re.compile(r'\d+/\d+(第二次)?报\d+天冻结[,，]*')
ORDER_PAT = re.compile(r'^([ZL]\d{4}-\d{1,3}-[YN])')


def clean_au(text):
    """AU列清洗：待→欠，删冻结报告，拆分，去纯数字"""
    if not text:
        return []
    t = str(text).strip()
    t = t.replace('待', '欠')
    t = FREEZE_PAT.sub('', t)
    t = re.sub(r'(\d+)/(\d+)(?!报)', r'\1月\2日', t)
    parts = re.split(r'[、,，;\n/]+', t)
    parts = [p.replace('月', '/').replace('日', '') if re.match(r'\d+/\d+', p) else p for p in parts]
    result = []
    for p in parts:
        p = p.strip().strip('，').strip('、').strip()
        if not p:
            continue
        if re.match(r'^[\d\s/\-]+$', p):
            continue
        result.append(p)
    return result


def classify_seg(seg):
    """分类单个AU片段：procurement/production/ignore/other"""
    for kw in IGNORE_KW:
        if kw in seg:
            return 'ignore'
    if '网笼丝杆' in seg and '尼龙网' in seg:
        return 'procurement'
    for kw in PROC_KW:
        if kw in seg:
            return 'procurement'
    for kw in PROD_KW:
        if kw in seg:
            return 'production'
    return 'other'


def format_tracking(items):
    """格式化异常跟踪列：仅第一个加'欠'，其余顿号分隔"""
    if not items:
        return ''
    cleaned = []
    for item in items:
        item = item.strip()
        if item.startswith('欠'):
            item = item[1:]
        if item:
            cleaned.append(item)
    return '欠' + '、'.join(cleaned) if cleaned else ''


def classify_av(av):
    """根据AV列分类"""
    if not av:
        return 'none'
    av = str(av).strip()
    exclude = ['技术', '工艺', '业务', '销售']
    has_excl = any(k in av for k in exclude)
    has_proc = '采购' in av or '外购' in av
    prod_kws = ['下料', '焊接', '组装', '粘网', '机加工', '机加', '喷砂', '喷粉', '外协']
    has_prod = any(k in av for k in prod_kws)
    if has_excl and not has_proc and not has_prod:
        return 'ignore'
    if has_proc:
        return 'procurement'
    if has_prod:
        return 'production'
    return 'other'


def fmt_date(val):
    """日期格式化为M/D"""
    if not val:
        return ''
    if isinstance(val, (datetime, date)):
        return f"{val.month}/{val.day}"
    return str(val).split(' ')[0]


def get_sheet_name(wb, prefix):
    """动态获取以prefix开头的Sheet名"""
    for name in wb.sheetnames:
        if name.startswith(prefix):
            return name
    return wb.sheetnames[0]


def get_delay(ws, row, category, av_text, today):
    """计算延期天数"""
    col_map = {'procurement': 16, '下料': 18, '机加': 30, '机加工': 30, '焊接': 34, '粘网': 36, '喷砂': 38, '喷粉': 40, '组装': 44}
    if category == 'procurement':
        col = 16
    else:
        col = None
        for kw in ['下料', '机加', '焊接', '粘网', '喷砂', '喷粉', '组装']:
            if kw in str(av_text or ''):
                col = col_map.get(kw)
                break
        if not col:
            col = 44
    val = ws.cell(row, col).value
    if not val:
        return None
    if isinstance(val, datetime):
        val = val.date()
    elif isinstance(val, date):
        pass
    else:
        try:
            val = datetime.strptime(str(val).split(' ')[0], '%Y-%m-%d').date()
        except:
            return None
    return (today - val).days


def get_plan_date(ws, row, category, av_text):
    """获取计划完成日期"""
    col_map = {'procurement': 16, '下料': 18, '机加': 30, '机加工': 30, '焊接': 34, '粘网': 36, '喷砂': 38, '喷粉': 40, '组装': 44}
    if category == 'procurement':
        col = 16
    else:
        col = None
        for kw in ['下料', '机加', '焊接', '粘网', '喷砂', '喷粉', '组装']:
            if kw in str(av_text or ''):
                col = col_map.get(kw)
                break
        if not col:
            col = 44
    val = ws.cell(row, col).value
    return fmt_date(val)


def get_actual(ws, row, av_text):
    """获取实际完成日期"""
    actual_map = {'下料': 19, '机加': 31, '机加工': 31, '焊接': 35, '粘网': 37, '喷砂': 39, '喷粉': 41, '组装': 45}
    for kw in ['下料', '机加', '焊接', '粘网', '喷砂', '喷粉', '组装']:
        if kw in str(av_text or ''):
            col = actual_map.get(kw)
            if col:
                return fmt_date(ws.cell(row, col).value)
            break
    return ''


def process_main_sheet(ws_main, today):
    """解析主计划表异常行"""
    anomaly_rows = []
    for r in range(5, ws_main.max_row + 1):
        au = ws_main.cell(r, 47).value
        av = ws_main.cell(r, 48).value
        if not au and not av:
            continue
        au_str = str(au or '')
        if any(kw in au_str for kw in ['已完成', '已完工', '已发货', '1.22号已发货']):
            if not av or str(av).strip() in ['None', '']:
                continue
        anomaly_rows.append({
            'row': r,
            'order': str(ws_main.cell(r, 2).value or ''),
            'category': str(ws_main.cell(r, 3).value or ''),
            'salesperson': str(ws_main.cell(r, 4).value or ''),
            'destination': str(ws_main.cell(r, 5).value or ''),
            'machine': str(ws_main.cell(r, 6).value or ''),
            'order_date': ws_main.cell(r, 7).value,
            'demand_date': ws_main.cell(r, 8).value,
            'quantity': ws_main.cell(r, 9).value,
            'au_raw': au_str,
            'av_raw': str(av or ''),
        })
    return anomaly_rows


def process_purchase_sheet(ws_pur):
    """解析外购物料表"""
    purchase_records = {}
    safety_stock_rows = []
    for r in range(2, ws_pur.max_row + 1):
        c_val = str(ws_pur.cell(r, 3).value or '').strip()
        d_val = ws_pur.cell(r, 4).value
        h_val = ws_pur.cell(r, 8).value
        i_val = ws_pur.cell(r, 9).value
        j_val = ws_pur.cell(r, 10).value
        k_val = ws_pur.cell(r, 11).value
        l_val = ws_pur.cell(r, 12).value
        m_val = ws_pur.cell(r, 13).value
        order_match = ORDER_PAT.match(c_val)
        if order_match:
            key = order_match.group(1)
            if key not in purchase_records:
                purchase_records[key] = []
            purchase_records[key].append({
                'row': r, 'task_no': c_val, 'material': str(d_val or ''),
                'purchase_qty': h_val, 'delivered_qty': i_val, 'owed_qty': j_val,
                'due_date': k_val, 'order_date': l_val, 'creator': str(m_val or '')})
        else:
            owed = j_val if j_val else 0
            try:
                owed = float(owed)
            except:
                owed = 0
            if owed > 0:
                safety_stock_rows.append({
                    'row': r, 'task_no': c_val, 'material': str(d_val or ''),
                    'purchase_qty': h_val, 'delivered_qty': i_val, 'owed_qty': j_val,
                    'due_date': k_val, 'order_date': l_val, 'creator': str(m_val or '')})
    return purchase_records, safety_stock_rows


def classify_anomalies(anomaly_rows, ws_main, today):
    """分类处理异常行"""
    prod_list, proc_list, ignore_list = [], [], []

    for item in anomaly_rows:
        av = item['av_raw']
        au = item['au_raw']
        av_cat = classify_av(av)
        segments = clean_au(au)

        prod_items, proc_items, ignore_items, other_items = [], [], [], []
        for seg in segments:
            t = classify_seg(seg)
            if t == 'procurement':
                proc_items.append(seg)
            elif t == 'production':
                prod_items.append(seg)
            elif t == 'ignore':
                ignore_items.append(seg)
            else:
                other_items.append(seg)

        if av_cat == 'procurement':
            final = 'procurement'
        elif av_cat == 'production':
            final = 'production'
        elif av_cat == 'ignore':
            final = 'ignore'
        else:
            if proc_items and not prod_items:
                final = 'procurement'
            elif prod_items and not proc_items:
                final = 'production'
            elif proc_items and prod_items:
                final = 'procurement'
            else:
                final = 'other'

        if final == 'production':
            tracking = format_tracking(prod_items)
            if not tracking and other_items:
                tracking = format_tracking(other_items)
            if not tracking:
                tracking = au.strip()
            first_resp = ''
            if '焊接' in av:
                first_resp = '张浩勋'
            elif any(kw in av for kw in ['下料', '粘网', '组装', '机加工', '机加']):
                first_resp = '郭先科'
            delay = get_delay(ws_main, item['row'], 'production', av, today)
            actual = get_actual(ws_main, item['row'], av)
            plan = get_plan_date(ws_main, item['row'], 'production', av)
            prod_list.append({
                'order': item['order'], 'category': item['category'], 'salesperson': item['salesperson'],
                'destination': item['destination'], 'machine': item['machine'],
                'order_date': fmt_date(item['order_date']), 'demand_date': fmt_date(item['demand_date']),
                'quantity': item['quantity'], 'plan_date': plan, 'actual_date': actual,
                'delay_days': delay, 'tracking': tracking, 'first_resp': first_resp,
                'next_resp': '', 'remark': '、'.join(other_items) if other_items else ''})

        elif final == 'procurement':
            tracking = format_tracking(proc_items)
            if not tracking and prod_items:
                tracking = format_tracking(prod_items)
            if not tracking:
                tracking = au.strip()
            delay = get_delay(ws_main, item['row'], 'procurement', av, today)
            plan = get_plan_date(ws_main, item['row'], 'procurement', av)
            proc_list.append({
                'order': item['order'], 'category': item['category'], 'salesperson': item['salesperson'],
                'destination': item['destination'], 'machine': item['machine'],
                'order_date': fmt_date(item['order_date']), 'demand_date': fmt_date(item['demand_date']),
                'quantity': item['quantity'], 'plan_date': plan, 'actual_date': '',
                'delay_days': delay, 'tracking': tracking, 'purchaser': '丰红伟',
                'remark': '、'.join(prod_items + other_items) if (prod_items or other_items) else ''})

        elif final == 'ignore':
            ignore_list.append({
                'order': item['order'], 'category': item['category'], 'salesperson': item['salesperson'],
                'machine': item['machine'], 'au_raw': au, 'av_raw': av,
                'reason': '、'.join(ignore_items) if ignore_items else '忽略项'})

    return prod_list, proc_list, ignore_list


def write_sheet(ws, title, headers, data, left_cols=None, today_str=''):
    """写入异常报表Sheet"""
    if left_cols is None:
        left_cols = set()
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    hdr_fill = PatternFill(start_color='B6DEE8', end_color='B6DEE8', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    yel_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    hdr_font = Font(name='宋体', size=12, bold=True)
    body_font = Font(name='宋体', size=12)
    title_font = Font(name='宋体', size=18, bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(1, 1, title)
    c.font = title_font
    c.alignment = center
    ws.row_dimensions[1].height = 30
    for i, h in enumerate(headers, 1):
        c = ws.cell(2, i, h)
        c.font = hdr_font
        c.alignment = center
        c.fill = hdr_fill
        c.border = thin
    ws.row_dimensions[2].height = 28
    for r_idx, row in enumerate(data, 3):
        ws.row_dimensions[r_idx].height = 23
        for c_idx, h in enumerate(headers, 1):
            val = row.get(h, '')
            c = ws.cell(r_idx, c_idx, val)
            c.font = body_font
            c.border = thin
            c.alignment = left if h in left_cols else center
            if h == '延期天数':
                d = row.get('_delay')
                if d is not None and d > 0:
                    c.fill = red_fill
                    c.font = Font(name='宋体', size=12, color='FFFFFF')
                elif d is not None and -3 <= d <= -1:
                    c.fill = yel_fill
    widths = {'序号': 6, '订单编号': 18, '类别': 8, '业务员': 8, '到货': 10, '机型': 18,
              '订单需求': 10, '订单交期': 10, '要求完成': 10, '实际完成': 10, '延期天数': 10,
              '异常跟踪': 42, '第一责任人': 10, '下工序责任人': 10, '采购人': 10, '备注': 35,
              'AU原文': 40, 'AV原文': 15, '忽略原因': 15}
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 12)


def generate_report(prod_list, proc_list, ignore_list, output_dir, date_label):
    """生成异常报表"""
    wb_rpt = openpyxl.Workbook()
    # 生产异常
    ws1 = wb_rpt.active
    ws1.title = f'{date_label}生产异常'
    ph = ['序号', '订单编号', '类别', '业务员', '到货', '机型', '订单需求', '订单交期', '要求完成', '实际完成', '延期天数', '异常跟踪', '第一责任人', '下工序责任人', '备注']
    pd = []
    for i, a in enumerate(prod_list, 1):
        d = a['delay_days']
        pd.append({'序号': i, '订单编号': a['order'], '类别': a['category'], '业务员': a['salesperson'],
            '到货': a['destination'], '机型': a['machine'], '订单需求': a['order_date'],
            '订单交期': a['demand_date'], '要求完成': a['plan_date'], '实际完成': a['actual_date'],
            '延期天数': str(d) if d is not None else '', '_delay': d,
            '异常跟踪': a['tracking'], '第一责任人': a['first_resp'], '下工序责任人': a['next_resp'], '备注': a['remark']})
    write_sheet(ws1, f'生产异常报表 {date_label}', ph, pd, {'异常跟踪', '备注', '机型'})
    # 采购异常
    ws2 = wb_rpt.create_sheet(f'{date_label}采购异常')
    uh = ['序号', '订单编号', '类别', '业务员', '到货', '机型', '订单需求', '订单交期', '要求完成', '实际完成', '延期天数', '异常跟踪', '采购人', '备注']
    ud = []
    for i, a in enumerate(proc_list, 1):
        d = a['delay_days']
        ud.append({'序号': i, '订单编号': a['order'], '类别': a['category'], '业务员': a['salesperson'],
            '到货': a['destination'], '机型': a['machine'], '订单需求': a['order_date'],
            '订单交期': a['demand_date'], '要求完成': a['plan_date'], '实际完成': a['actual_date'],
            '延期天数': str(d) if d is not None else '', '_delay': d,
            '异常跟踪': a['tracking'], '采购人': a['purchaser'], '备注': a['remark']})
    write_sheet(ws2, f'采购异常报表 {date_label}', uh, ud, {'异常跟踪', '备注', '机型'})
    # 忽略项
    ws3 = wb_rpt.create_sheet(f'{date_label}忽略项')
    ih = ['序号', '订单编号', '类别', '业务员', '机型', 'AU原文', 'AV原文', '忽略原因']
    id_ = []
    for i, a in enumerate(ignore_list, 1):
        id_.append({'序号': i, '订单编号': a['order'], '类别': a['category'], '业务员': a['salesperson'],
            '机型': a['machine'], 'AU原文': a['au_raw'], 'AV原文': a['av_raw'], '忽略原因': a['reason']})
    write_sheet(ws3, f'忽略项 {date_label}', ih, id_, {'AU原文', 'AV原文', '忽略原因', '机型'})
    rpt_path = os.path.join(output_dir, f'伟良异常报表{date_label}.xlsx')
    wb_rpt.save(rpt_path)
    return rpt_path


def generate_safety_stock(safety_stock_rows, ws_pur, output_dir, date_label):
    """生成安全库存未到货表"""
    wb_s = openpyxl.Workbook()
    ws_s = wb_s.active
    ws_s.title = f'{date_label}安全库存未到货'
    sh = ['供应商', '采购单号', '任务单号', '物料名称', '图号', '基本单位', '收料数量', '采购数量', '已交货数量', '欠交货数量', '要求交期', '下单日期', '创建人', '未税金额', '备注']
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    hdr_fill = PatternFill(start_color='B6DEE8', end_color='B6DEE8', fill_type='solid')
    hdr_font = Font(name='宋体', size=12, bold=True)
    body_font = Font(name='宋体', size=12)
    title_font = Font(name='宋体', size=18, bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_s.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(sh))
    c = ws_s.cell(1, 1, f'安全库存未到货 {date_label}')
    c.font = title_font
    c.alignment = center
    ws_s.row_dimensions[1].height = 30
    for i, h in enumerate(sh, 1):
        c = ws_s.cell(2, i, h)
        c.font = hdr_font
        c.alignment = center
        c.fill = hdr_fill
        c.border = thin
    ws_s.row_dimensions[2].height = 28
    for r_idx, item in enumerate(safety_stock_rows, 3):
        ws_s.row_dimensions[r_idx].height = 23
        orig = item['row']
        for c_idx in range(1, len(sh) + 1):
            val = ws_pur.cell(orig, c_idx).value
            c = ws_s.cell(r_idx, c_idx, val)
            c.font = body_font
            c.border = thin
            c.alignment = center
        for dc in [11, 12]:
            v = ws_s.cell(r_idx, dc).value
            if v and isinstance(v, (datetime, date)):
                ws_s.cell(r_idx, dc).value = v.strftime('%Y/%m/%d')
    for i, w in enumerate([25, 16, 18, 20, 20, 8, 10, 10, 10, 10, 14, 14, 8, 12, 20], 1):
        ws_s.column_dimensions[get_column_letter(i)].width = w
    safe_path = os.path.join(output_dir, f'安全库存未到货{date_label}.xlsx')
    wb_s.save(safe_path)
    return safe_path


def process_finished_goods(fg_file, main_file, output_dir, date_label):
    """成品入库表匹配主计划表，涂绿色标记"""
    wb_fg = openpyxl.load_workbook(fg_file, data_only=True)
    ws_fg = wb_fg[get_sheet_name(wb_fg, 'Sheet')]
    # 收集入库订单号
    fg_orders = set()
    for r in range(2, ws_fg.max_row + 1):
        order = str(ws_fg.cell(r, 3).value or '').strip()
        if order:
            # 精确匹配 + 前缀匹配
            m = ORDER_PAT.match(order)
            if m:
                fg_orders.add(m.group(1))
            fg_orders.add(order)
    # 打开主计划表
    wb_main = openpyxl.load_workbook(main_file)
    ws_main = wb_main[get_sheet_name(wb_main, '订单管理表')]
    green_fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
    matched = 0
    for r in range(5, ws_main.max_row + 1):
        cell = ws_main.cell(r, 2)
        order = str(cell.value or '').strip()
        if order and order in fg_orders:
            cell.fill = green_fill
            matched += 1
    # 取消所有隐藏行
    for r in range(1, ws_main.max_row + 1):
        ws_main.row_dimensions[r].hidden = False
    fg_path = os.path.join(output_dir, f'主计划表{date_label}-成品入库标记.xlsx')
    wb_main.save(fg_path)
    return fg_path, matched


def main():
    parser = argparse.ArgumentParser(description='伟良日报处理')
    parser.add_argument('--date', required=True, help='日期 YYYY-MM-DD')
    parser.add_argument('--main-file', required=True, help='主计划表路径')
    parser.add_argument('--purchase-file', required=True, help='外购物料表路径')
    parser.add_argument('--output-dir', required=True, help='输出目录')
    parser.add_argument('--finished-goods-file', help='成品入库表路径（可选）')
    args = parser.parse_args()

    today = datetime.strptime(args.date, '%Y-%m-%d').date()
    date_label = f'{today.month}.{today.day}'
    os.makedirs(args.output_dir, exist_ok=True)

    print("=" * 60)
    print(f"伟良日报处理 - {date_label}")
    print("=" * 60)

    # 加载文件
    wb_main = openpyxl.load_workbook(args.main_file, data_only=True)
    main_sheet_name = get_sheet_name(wb_main, '订单管理表')
    ws_main = wb_main[main_sheet_name]
    wb_pur = openpyxl.load_workbook(args.purchase_file, data_only=True)
    ws_pur = wb_pur[get_sheet_name(wb_pur, 'Sheet')]

    # Step1: 解析主计划表
    print("\n[1] 解析主计划表异常行...")
    anomaly_rows = process_main_sheet(ws_main, today)
    print(f"  找到 {len(anomaly_rows)} 个异常行")

    # Step2: 解析外购物料表
    print("\n[2] 解析外购物料到货跟踪表...")
    purchase_records, safety_stock_rows = process_purchase_sheet(ws_pur)
    print(f"  订单号: {len(purchase_records)}个, {sum(len(v) for v in purchase_records.values())}行")
    print(f"  安全库存未到货: {len(safety_stock_rows)}行")

    # Step3: 分类处理
    print("\n[3] 分类处理...")
    prod_list, proc_list, ignore_list = classify_anomalies(anomaly_rows, ws_main, today)
    print(f"  生产异常: {len(prod_list)}行")
    print(f"  采购异常: {len(proc_list)}行")
    print(f"  忽略项: {len(ignore_list)}行")

    # Step4: 生成异常报表
    print("\n[4] 生成异常报表...")
    rpt_path = generate_report(prod_list, proc_list, ignore_list, args.output_dir, date_label)
    print(f"  异常报表: {rpt_path}")

    # Step5: 安全库存
    print("\n[5] 生成安全库存未到货表...")
    safe_path = generate_safety_stock(safety_stock_rows, ws_pur, args.output_dir, date_label)
    print(f"  安全库存: {safe_path}")

    # Step6: 成品入库标记（可选）
    fg_path = None
    fg_matched = 0
    if args.finished_goods_file:
        print("\n[6] 成品入库表匹配...")
        fg_path, fg_matched = process_finished_goods(args.finished_goods_file, args.main_file, args.output_dir, date_label)
        print(f"  匹配 {fg_matched} 个订单涂绿色: {fg_path}")

    # 汇总
    print("\n" + "=" * 60)
    print("处理完成")
    print("=" * 60)
    print(f"生产异常: {len(prod_list)}行 | 采购异常: {len(proc_list)}行 | 忽略项: {len(ignore_list)}行 | 安全库存未到货: {len(safety_stock_rows)}行")
    if fg_path:
        print(f"成品入库匹配: {fg_matched}个订单")
    print(f"\n产物:")
    print(f"  1. {rpt_path}")
    print(f"  2. {safe_path}")
    if fg_path:
        print(f"  3. {fg_path}")


if __name__ == '__main__':
    main()
